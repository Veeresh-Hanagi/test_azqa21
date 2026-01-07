import unittest
import time
from unittest import TextTestRunner, expectedFailure

import faker
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

fake = faker.Faker()

# -------------------------------
# Create Excel Test Data
# -------------------------------
wb = Workbook()
sheet = wb.active
sheet.title = "LoginData"

# Generate fake email/password
for i in range(1, 20):
    sheet.cell(row=i, column=1).value = fake.email()
    sheet.cell(row=i, column=2).value = fake.password()

# Add few known test users
sheet.cell(row=4, column=1).value = "welldocsu"
sheet.cell(row=8, column=1).value = "welldocsu"
sheet.cell(row=12, column=1).value = "welldocsu"
sheet.cell(row=16, column=1).value = "welldocsu"
sheet.cell(row=4, column=2).value = "welldoc123"
sheet.cell(row=8, column=2).value = "welldoc123"
sheet.cell(row=12, column=2).value = "welldoc123"
sheet.cell(row=16, column=2).value = "welldoc1234"

wb.save("TestData.xlsx")
wb.close()

# Reopen for reading/writing
wb = load_workbook("TestData.xlsx")
sheet = wb.active
rows = sheet.max_row
cols = sheet.max_column

print(f"Total Rows: {rows}, Columns: {cols}")

# -------------------------------
# Define Fill Colors
# -------------------------------
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# -------------------------------
# Test Class
# -------------------------------
class LoginTest(unittest.TestCase):
    successful_users = []  # ✅ shared class variable for successful logins

    def setUp(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get("https://azqa21-dsm.testwd.com/SMITPortal/Guest/Login.htm")
        self.wait = WebDriverWait(self.driver, 10)

    def test_01_login(self,):
        """✅ Test to validate login and store successful credentials"""
        driver = self.driver
        wait = self.wait

        for i in range(1, rows + 1):
            try:
                # Wait for username/password fields
                username_field = wait.until(EC.presence_of_element_located((By.ID, "cmusername")))
                password_field = wait.until(EC.presence_of_element_located((By.ID, "cmpassword")))
                login_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@data-localize='Login.LoginBtn']")))

                # Clear fields
                username_field.clear()
                password_field.clear()

                # Read data from Excel
                username = sheet.cell(row=i, column=1).value
                password = sheet.cell(row=i, column=2).value

                print(f"Attempting login with: {username} / {password}")

                # Send keys
                username_field.send_keys(username)
                password_field.send_keys(password)
                login_button.click()

                # Wait briefly for potential result
                time.sleep(2)

                # Check for invalid message
                try:
                    invalid_text_elem = wait.until(
                        EC.presence_of_element_located((By.XPATH, "//div[@role='alert']"))
                    )
                    invalid_message = invalid_text_elem.text.strip()

                    if invalid_message == "The username or password you entered is incorrect. Please try again.":
                        print(f"[{i}] ❌ Login failed.")
                        sheet.cell(row=i, column=3).value = "Failed"
                        sheet.cell(row=i, column=3).fill = red_fill
                    else:
                        print(f"[{i}] ✅ Login passed.")
                        sheet.cell(row=i, column=3).value = "Passed"
                        sheet.cell(row=i, column=3).fill = green_fill
                        # ✅ Store successful credentials
                        LoginTest.successful_users.append((username, password))

                except TimeoutException:
                    # If no error alert appeared, assume login success
                    print(f"[{i}] ✅ Login successful (no error message shown).")
                    sheet.cell(row=i, column=3).value = "Passed"
                    sheet.cell(row=i, column=3).fill = green_fill
                    # ✅ Store successful credentials
                    LoginTest.successful_users.append((username, password))

                # Save after each iteration
                wb.save("TestData.xlsx")

                # Navigate back for next iteration
                driver.get("https://azqa21-dsm.testwd.com/SMITPortal/Guest/Login.htm")

            except (NoSuchElementException, TimeoutException) as e:
                print(f"[{i}] ❌ Exception during login: {e}")
                sheet.cell(row=i, column=3).value = "Failed"
                sheet.cell(row=i, column=3).fill = red_fill
                wb.save("TestData.xlsx")
                driver.get("https://azqa21-dsm.testwd.com/SMITPortal/Guest/Login.htm")

    def test_02_Patientoverview(self):
        """✅ Use successful login credentials from test_login"""
        driver = self.driver
        wait = self.wait

        if not LoginTest.successful_users:
            self.skipTest("No successful login credentials found from previous test.")

        # Use the first successful credential
        username, password = LoginTest.successful_users[0]
        print(f"\nReusing successful credentials: {username} / {password}")

        # Login again using successful credentials
        username_field = wait.until(EC.presence_of_element_located((By.ID, "cmusername")))
        password_field = wait.until(EC.presence_of_element_located((By.ID, "cmpassword")))
        login_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@data-localize='Login.LoginBtn']")))

        username_field.clear()
        password_field.clear()
        username_field.send_keys(username)
        password_field.send_keys(password)
        login_button.click()

        # Wait for dashboard/logbook page to load
        time.sleep(3)

        try:
            # ✅ Corrected locator syntax for Patient List tab
            patientlist_tab = wait.until(
                EC.presence_of_element_located((By.XPATH, "//span[@data-localize='Main.PatientList']"))
            )
            print("✅ Patient List tab found")
        except TimeoutException:
            print("❌ Patient List tab not found after login.")

        try:
            # ✅ Corrected locator syntax for Name search input
            name_search_input = wait.until(
                EC.presence_of_element_located((By.XPATH, "//input[@id='NameSearch']"))
            )
            print("✅ Name search input found")
            # Example action
            name_search_input.clear()
            name_search_input.send_keys("tdc387")
        except TimeoutException:
            print("❌ Name search input not found")
        self.search_btn = driver.find_element(By.XPATH,"//button[@class='btn btn-default search-btn-spacing searchPatient']//span[@data-localize='Main.Search'][normalize-space()='Search']")
        self.search_btn.click()

        try:
            # ✅ Corrected locator syntax for Patient
            patient= wait.until(
                EC.presence_of_element_located((By.XPATH, "//td[@class='patient']"))
            )
            patient.click()
            time.sleep(5)
        except TimeoutException:
            print("❌ patient not found")

        try:
            # ✅ Corrected locator syntax for patient logbook
            patient_logbook= wait.until(
                EC.presence_of_element_located((By.XPATH, "//li[@class='mr-3']//a[normalize-space()='Logbook']"))
            )
            patient_logbook.click()
            time.sleep(5)
        except TimeoutException:
            print("patient logbook not found")
        try:
            # ✅ Corrected locator syntax for patient Medications
            patient_Medications = wait.until(
                EC.presence_of_element_located((By.XPATH, "//a[normalize-space()='Medications']"))
            )
            patient_Medications.click()
            time.sleep(5)
        except TimeoutException:
            print("patient Medications not found")
        try:
            # ✅ Corrected locator syntax for patient Health info
            patient_healthinfo = wait.until(
                EC.presence_of_element_located((By.XPATH, "//li[@aria-label='Health Info ']"))
            )
            patient_healthinfo.click()
            time.sleep(5)
        except TimeoutException:
            print("patient healthinfo not found")
        try:
            # ✅ Corrected locator syntax for patient curriculam
            patient_curriculm = wait.until(
                EC.presence_of_element_located((By.XPATH, "//a[normalize-space()='Curriculum']"))
            )
            patient_curriculm.click()
            time.sleep(5)
        except TimeoutException:
            print("patient curriculm not found")
        try:
            # ✅ Corrected locator syntax for patient program status
            patient_ProgramStatus = wait.until(
                EC.presence_of_element_located((By.XPATH, "//a[normalize-space()='Program Status']"))
            )
            patient_ProgramStatus.click()
            time.sleep(5)
        except TimeoutException:
            print("patient program status not found")

    def test_03_logbook(self):
        driver = self.driver
        wait = self.wait

        if not LoginTest.successful_users:
            self.skipTest("No successful login credentials found from previous test.")

        # Use the first successful credential
        username, password = LoginTest.successful_users[0]
        print(f"\nReusing successful credentials: {username} / {password}")

        # Login again using successful credentials
        username_field = wait.until(EC.presence_of_element_located((By.ID, "cmusername")))
        password_field = wait.until(EC.presence_of_element_located((By.ID, "cmpassword")))
        login_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@data-localize='Login.LoginBtn']")))

        username_field.clear()
        password_field.clear()
        username_field.send_keys(username)
        password_field.send_keys(password)
        login_button.click()

        # Wait for dashboard/logbook page to load
        time.sleep(3)
        try:
            # ✅ Corrected locator syntax for Patient List tab
            patientlist_tab = wait.until(
                EC.presence_of_element_located((By.XPATH, "//span[@data-localize='Main.PatientList']"))
            )
            print("✅ Patient List tab found")
        except TimeoutException:
            print("❌ Patient List tab not found after login.")

        try:
            # ✅ Corrected locator syntax for Name search input
            name_search_input = wait.until(
                EC.presence_of_element_located((By.XPATH, "//input[@id='NameSearch']"))
            )
            print("✅ Name search input found")
            # Example action
            name_search_input.clear()
            name_search_input.send_keys("aderx161")
        except TimeoutException:
            print("❌ Name search input not found")
        self.search_btn = driver.find_element(By.XPATH,"//button[@class='btn btn-default search-btn-spacing searchPatient']//span[@data-localize='Main.Search'][normalize-space()='Search']")
        self.search_btn.click()

        try:
            # ✅ Corrected locator syntax for Patient
            patient= wait.until(
                EC.presence_of_element_located((By.XPATH, "//td[@class='patient']"))
            )
            patient.click()
            time.sleep(5)
        except TimeoutException:
            print("❌ patient not found")

        try:
            # ✅ Corrected locator syntax for patient logbook
            patient_logbook= wait.until(
                EC.presence_of_element_located((By.XPATH, "//li[@class='mr-3']//a[normalize-space()='Logbook']"))
            )
            patient_logbook.click()
            time.sleep(5)
        except TimeoutException:
            print("patient logbook not found")

        # self.element= driver.find_element(By.XPATH, "//span[@aria-label='calender datepicker modal dialog']")
        # self.driver.execute_script("arguments[0].scrollIntoView();", self.element)

        try:
            # ✅ Corrected locator syntax for patient logbook datepicker
            patient_logbook_datepicker= wait.until(
                EC.presence_of_element_located((By.XPATH, "//span[@aria-label='calender datepicker modal dialog']"))
            )
            patient_logbook_datepicker.click()
        except TimeoutException:
            print("patient logbook  datepicker is not found")

        try:
            # ✅ Corrected locator syntax for patient logbook datepicker fromdate
            logbook_datepicker_fromdate= wait.until(
                EC.presence_of_element_located((By.XPATH, "//div[@class='LogbookAccordContent']//img[@title='Calendar View']"))
            )
            logbook_datepicker_fromdate.click()
        except TimeoutException:
            print("patient logbook  datepicker fromdate picker  is not found")

        while True :
            day="9"
            month ="August"
            year="2025"

            yrr=driver.find_element(By.XPATH,"//span[@class='ui-datepicker-year']").text
            mon=driver.find_element(By.XPATH,"//span[@class='ui-datepicker-month']").text

            if yrr==year and mon==month:
                break
            else:
                driver.find_element(By.XPATH,"//span[@class='ui-icon ui-icon-circle-triangle-w']").click()

        dates=driver.find_elements(By.XPATH,"//table[@class='ui-datepicker-calendar']/tbody/tr/td")

        for date in dates:
            if date.text==day:
                date.click()
        time.sleep(10)
        try:
            # ✅ Corrected locator syntax for patient logbook datepicker Todate
            logbook_datepicker_Todate= wait.until(
                EC.presence_of_element_located((By.XPATH, "//div[@class='LogbookAccordContent DateRangeSpace']//img[@title='Calendar View']"))
            )
            logbook_datepicker_Todate.click()
        except TimeoutException:
            print("logbook datepicker todate is not found")
        while True:
                day = "10"
                month = "October"
                year = "2025"

                yrr = driver.find_element(By.XPATH, "//span[@class='ui-datepicker-year']").text
                mon = driver.find_element(By.XPATH, "//span[@class='ui-datepicker-month']").text

                if yrr == year and mon == month:
                    break
                else:
                    driver.find_element(By.XPATH, "//span[@class='ui-icon ui-icon-circle-triangle-w']").click()

        dates = driver.find_elements(By.XPATH, "//table[@class='ui-datepicker-calendar']/tbody/tr/td")

        for date in dates:
            if date.text == day:
              date.click()
        time.sleep(10)
        self.driver.find_element(By.XPATH,"//button[@id='SetdateRangeButton']").click()
        time.sleep(5)







    def tearDown(self):
        time.sleep(2)
        self.driver.quit()

if __name__ == "__main__":
    suite = unittest.TestSuite()
    suite.addTest(LoginTest("test_01_login"))
    suite.addTest(LoginTest("test_02_Patientoverview"))
    suite.addTest(LoginTest("test_03_logbook"))
    runner = TextTestRunner()
    runner.run(suite)
